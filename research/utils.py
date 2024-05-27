import pandas as pd
import numpy as np
from openpyxl import load_workbook
import xlrd
import xlwt
from pandas.io.excel import ExcelWriter
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from pyexcelerate import Workbook
import os

def extract_text_from_txt(txt_path):
    with open(txt_path, 'r', encoding='utf-8') as file:
        text = file.read()
    return text.replace("※", "")

def convertXls2XlsX(file_path: str, filename_suffix = ''):
    if not file_path.endswith('.xls'):
        raise ValueError("The file path must end with '.xls'")
    
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names
    
    df = pd.read_excel(file_path, sheet_name=sheet_names[0])
    xlsx_file_path = file_path.replace('.xls', f"_{filename_suffix}.xlsx")
    
    with pd.ExcelWriter(xlsx_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_names[0])
        
    return xlsx_file_path

def convertXlsx2Xls(file_path: str, filename_suffix = ''):
    if not file_path.endswith('.xlsx'):
        raise ValueError("The file path must end with '.xlsx'")

    xlsx = pd.ExcelFile(file_path)
    sheet_names = xlsx.sheet_names
    
    df = pd.read_excel(file_path, sheet_name=sheet_names[0])
    xls_file_path = file_path.replace('.xlsx', f"_{filename_suffix}.xls")

    with pd.ExcelWriter(xls_file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_names[0])
        
        # worksheet = writer.sheets[sheet_names[0]]
        
        # column_widths = {
        #     'jp': 30,
        #     'vn (original)': 30,
        #     'vn (updated)': 40,
        #     'Reason': 40,
        #     'en': 40,
        #     'zh-tc': 40
        # }

        # format_worksheet(df, worksheet, column_widths)

    return xls_file_path

def format_worksheet(df, worksheet, column_widths):
    for col_name, col_width in column_widths.items():
            col_idx = df.columns.get_loc(col_name) + 1
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = col_width
            
    # Adjust row heights by contents of cell
    def adjust_row_height(worksheet):
        for row in worksheet.iter_rows():
            max_height = 0
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if cell.value:
                    cell_length = len(str(cell.value))
                    column_width = worksheet.column_dimensions[get_column_letter(cell.column)].width
                    # Calculate the height based on the content length and column width
                    height = (cell_length // column_width + 1) * 15  # Assuming 15 points per line
                    max_height = max(max_height, height)
            worksheet.row_dimensions[row[0].row].height = max_height

    adjust_row_height(worksheet)

def add_new_sub_dir_file_path(file_path, new_subdirectory = 'updated_outputs'):
    directory, file_name = os.path.split(file_path)
    
    new_directory = os.path.join(directory, new_subdirectory)
    
    os.makedirs(new_directory, exist_ok=True)
    
    new_file_path = os.path.join(new_directory, file_name)
    
    return new_file_path

def add_suffix_to_filename(file_path, suffix):
    directory, filename = os.path.split(file_path)
    
    name, ext = os.path.splitext(filename)
    
    new_filename = f"{name}_{suffix}{ext}"
    
    new_file_path = os.path.join(directory, new_filename)
    
    return new_file_path